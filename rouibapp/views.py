from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse

from .forms import EmployeForm
from .models import ArretPET, CausesPET, Equipe, EquipePET, FactionPET, LignePET, MachinePET, Production, Ligne, Machine, Causes, ProductionPET, Responsable, ResponsablePET, Type, Arret,Employe
import json
from django.db import models


# ===== AUTHENTIFICATION =====
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Nom d’utilisateur ou mot de passe incorrect.')
    return render(request, 'rouibapp/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ===== PAGES =====



def equipe(request):
    equipe_carton = Equipe.objects.all()
    equipe_pet = EquipePET.objects.all()
    
    context = {
        'equipe_carton': equipe_carton,
        'equipe_pet': equipe_pet,
    }
    return render(request, 'rouibapp/equipe.html', context)



def production(request):
    return render(request, 'rouibapp/production.html')



def safe_int(val, default=0):
    try:
        return int(val)
    except (TypeError, ValueError):
        return default

def safe_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default




# ===== PAGE CARTON =====
def carton(request):
    productions_list = Production.objects.select_related('equipe', 'ligne').all().order_by('-date')
    paginator = Paginator(productions_list, 5)
    page_number = request.GET.get('page')
    productions = paginator.get_page(page_number)

    context = {
        'productions': productions,
        'lignes': Ligne.objects.all(),
        'responsables': Responsable.objects.all(),
        'arrets': Arret.objects.all(),
        'equipe': Equipe.objects.all(),
        'types': Type.objects.all(),
        'machines': Machine.objects.all(),
        'causes': Causes.objects.all(),
    }
    return render(request, 'rouibapp/carton.html', context)

def dm(request):
    from .models import Ligne, Arret, Production, Dm
    context = {
        'lignes': Ligne.objects.all(),
        'arrets': Arret.objects.all(),
        'productions': Production.objects.all(),
        'dms': Dm.objects.all(),
    }
    return render(request, 'rouibapp/dm.html', context)


# ===== AJOUT PRODUCTION =====
def safe_int(val, default=0):
    try:
        return int(val)
    except (TypeError, ValueError):
        return default

def safe_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default

def ajouter_production(request):
    if request.method == 'POST':
        date = request.POST.get('date')
        ligne_id = request.POST.get('ligne')
        equipe_id = request.POST.get('equipe')
        quart = request.POST.get('quart')
        
        # CORRECTION: Validation et conversion des champs numériques
        cadence = safe_float(request.POST.get('cadence', 0))
        temps_brut = safe_int(request.POST.get('temps_brut', 8))  # Valeur par défaut 8
        temps_programme = safe_int(request.POST.get('temps_programme', 0))
        theorique = safe_int(request.POST.get('theorique', 0))
        realise = safe_int(request.POST.get('realise', 0))
        C = safe_float(request.POST.get('C', 0))
        intrants = safe_float(request.POST.get('intrants', 0))
        technique = safe_float(request.POST.get('technique', 0))
        operationnel = safe_float(request.POST.get('operationnel', 0))
        siroperie = safe_float(request.POST.get('siroperie', 0))
        utilite = safe_float(request.POST.get('utilite', 0))
        espace = safe_float(request.POST.get('espace', 0))
        perte = safe_int(request.POST.get('perte', 0))
        
        # Validation des champs NEP et NIA
        nep_complet_mp_theorique = safe_int(request.POST.get('nep_complet_mp_theorique', 0))
        nep_complet_mp_reel = safe_float(request.POST.get('nep_complet_mp_reel', 0))
        nep_soude_mp_theorique = safe_float(request.POST.get('nep_soude_mp_theorique', 0))
        nep_soude_mp_reel = safe_float(request.POST.get('nep_soude_mp_reel', 0))
        nia_theorique = safe_float(request.POST.get('nia_theorique', 0))
        nia_reel = safe_int(request.POST.get('nia_reel', 0))
        revision_theorique = safe_float(request.POST.get('revision_theorique', 0))
        revision_reel = safe_float(request.POST.get('revision_reel', 0))

        ligne = get_object_or_404(Ligne, pk=ligne_id) if ligne_id else None
        equipe = get_object_or_404(Equipe, pk=equipe_id) if equipe_id else None

       


        
        # CALCULS AUTOMATIQUES (comme dans add_production)
        temps_non_programme = temps_brut - temps_programme

       
        
        # Recalculer le théorique si nécessaire
        if cadence > 0 and temps_programme > 0:
            theorique = cadence * temps_programme
        
        ecart = theorique - realise
        trs = (realise / theorique * 100) if theorique > 0 else 0
        temps_productif = (realise / cadence) if cadence > 0 else 0
        heure_arrets = (ecart / cadence) if cadence > 0 else 0
        
        # CALCULS DES ÉCARTS - C'EST ICI LE PROBLÈME !
        nep_complet_ecart = nep_complet_mp_reel - nep_complet_mp_theorique
        nep_soude_ecart = nep_soude_mp_reel - nep_soude_mp_theorique
        nia_ecart = nia_reel - nia_theorique
        
        # CALCULS DES TOTAUX
        total_arrets_theorique_program = nep_complet_mp_theorique + nep_soude_mp_theorique + nia_theorique
        total_arrets_reels_program = nep_complet_mp_reel + nep_soude_mp_reel + nia_reel
        
        # CALCUL DU TOTAL DÉPASSEMENTS - CORRECTION ICI !
        total_depass = nep_complet_ecart + nep_soude_ecart + nia_ecart
        
        revision_ecart = revision_reel - revision_theorique
        
        # CALCUL DES ARRÊTS NON PROGRAMMÉS - À ADAPTER SELON VOS BESOINS
        # Récupérer les arrêts de la base de données
        arret = Arret.objects.filter(
            date_arret=date,
            quart=quart
        ).first()
        
        if arret:
            total_arrets_minutes = safe_int(arret.duree_arret, 0)
            total_arrets_non_progr = round(total_arrets_minutes / 60, 3)  # Convertir en heures
        else:
            total_arrets_non_progr = 0
            
        total_arrets = total_arrets_reels_program + total_arrets_non_progr
        total = temps_productif + total_arrets
        
        Production.objects.create(
            date=date,
            ligne=ligne,
            equipe=equipe,
            quart=quart,
            cadence=cadence,
            temps_brut=temps_brut,
            temps_programme=temps_programme,
            temps_non_programme=temps_non_programme,
            theorique=theorique,
            realise=realise,
            trs=trs,
            C=C,
            intrants=intrants,
            technique=technique,
            operationnel=operationnel,
            siroperie=siroperie,
            utilite=utilite,
            espace=espace,
            perte=perte,
            ecart=ecart,
            temps_productif=temps_productif,
            heure_arrets=heure_arrets,
            nep_complet_mp_theorique=nep_complet_mp_theorique,
            nep_complet_mp_reel=nep_complet_mp_reel,
            nep_complet_ecart=nep_complet_ecart,
            nep_soude_mp_theorique=nep_soude_mp_theorique,
            nep_soude_mp_reel=nep_soude_mp_reel,
            nep_soude_ecart=nep_soude_ecart,
            nia_theorique=nia_theorique,
            nia_reel=nia_reel,
            nia_ecart=nia_ecart,
            total_arrets_theorique_program=total_arrets_theorique_program,
            total_arrets_reels_program=total_arrets_reels_program,
            total_depass=total_depass,  # ← MAINTENANT CALCULÉ !
            revision_theorique=revision_theorique,
            revision_reel=revision_reel,
            revision_ecart=revision_ecart,
            total_arrets_non_progr=total_arrets_non_progr,  # ← MAINTENANT CALCULÉ !
            total_arrets=total_arrets,
            total=total,
        )
        messages.success(request, "Production ajoutée avec succès.")
        return redirect('production')
    return redirect('production')


def modifier_production(request, id):
    production = get_object_or_404(Production, pk=id)
    
    if request.method == 'POST':
        # Récupération des données de base
        production.date = request.POST.get('date')
        ligne_id = request.POST.get('ligne')
        equipe_id = request.POST.get('equipe')
       
        # Mise à jour des relations
        production.ligne = get_object_or_404(Ligne, pk=ligne_id) if ligne_id else None
        production.equipe = get_object_or_404(Equipe, pk=equipe_id) if equipe_id else None

        # Récupérer les valeurs du formulaire
        cadence = float(request.POST.get('cadence', 0))
        temps_programme = float(request.POST.get('temps_programme', 0))
        realise = float(request.POST.get('realise', 0))
        temps_brut = float(request.POST.get('temps_brut', 8))
        temps_programme = float(request.POST.get('temps_programme', 0))
        

        # Recalculer les champs
        theorique = cadence * temps_programme
        ecart = theorique - realise
        trs = (realise / theorique * 100) if theorique > 0 else 0
        temps_productif = (realise / cadence) if cadence > 0 else 0
        heure_arrets = (ecart / cadence) if cadence > 0 else 0
        temps_non_programme = temps_brut - temps_programme
        production.temps_non_programme = temps_non_programme
       
        # Mise à jour des autres champs
        production.quart = request.POST.get('quart')
        production.cadence = cadence
        production.temps_programme = temps_programme
        production.realise = realise
        production.theorique = theorique
        production.ecart = ecart
        production.trs = trs
        production.temps_productif = temps_productif
        production.heure_arrets = heure_arrets
        production.C = request.POST.get('C')
        production.perte = request.POST.get('perte')
        
        # Champs NEP et NIA
        production.nep_complet_mp_theorique = request.POST.get('nep_complet_mp_theorique')
        production.nep_complet_mp_reel = request.POST.get('nep_complet_mp_reel')
        production.nep_soude_mp_theorique = request.POST.get('nep_soude_mp_theorique')
        production.nep_soude_mp_reel = request.POST.get('nep_soude_mp_reel')
        production.nia_theorique = request.POST.get('nia_theorique')
        production.nia_reel = request.POST.get('nia_reel')
        production.revision_theorique = request.POST.get('revision_theorique')
        production.revision_reel = request.POST.get('revision_reel')
        
        # Mise à jour du format_produit si fourni
        format_produit = request.POST.get('format_produit')
        if format_produit:
            production.format_produit = format_produit
        
        try:
            production.save()
            messages.success(request, "Production modifiée avec succès.")
            return redirect('production')
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification : {str(e)}")
            return redirect('production')
    
    return redirect('production')


def supprimer_production(request, id):
    production = get_object_or_404(Production, pk=id)
    production.delete()
    messages.success(request, "Production supprimée avec succès.")
    return redirect('production')



# ===== CORRECTION FONCTION AJOUTER EMPLOYE =====
def ajouter_employe(request):
    if request.method == 'POST':
        form = EmployeForm(request.POST)
        if form.is_valid():
            employe = form.save(commit=False)
            
            if employe.type_equipe == 'PET':
                # Ajout dans EquipePET
                EquipePET.objects.create(
                    nom=employe.nom,
                    prenom=employe.prenom
                )
                employe.save()
                messages.success(request, f'Employé {employe.prenom} {employe.nom} ajouté à l\'équipe PET avec succès!')
                return redirect('equipe')  # ✅ Rediriger vers la page équipe au lieu de PET
                
            else:  # type_equipe == 'CARTON'
                # Ajout dans Equipe "Carton"
                Equipe.objects.create(
                    nom=employe.nom,
                    prenom=employe.prenom
                )
                employe.save()
                messages.success(request, f'Employé {employe.prenom} {employe.nom} ajouté à l\'équipe Carton avec succès!')
                return redirect('equipe')  # ✅ Rediriger vers la page équipe au lieu de carton
        else:
            messages.error(request, 'Erreur dans le formulaire. Veuillez vérifier les données saisies.')
    else:
        form = EmployeForm()

    return render(request, 'rouibapp/employe_form.html', {'form': form})  



def modifier_employe(request, id_equipe):
    employe = get_object_or_404(Equipe, id_equipe=id_equipe)
    if request.method == 'POST':
        employe.nom = request.POST.get('nom')
        employe.prenom = request.POST.get('prenom')
        employe.save()
    return redirect('equipe')


def supprimer_employe(request, id_equipe):
    employe = get_object_or_404(Equipe, id_equipe=id_equipe)
    employe.delete()
    return redirect('equipe')

def modifier_employe_pet(request, id_equipe):
    """Vue pour modifier un employé PET"""
    employe = get_object_or_404(EquipePET, id_equipe=id_equipe)
    if request.method == 'POST':
        employe.nom = request.POST.get('nom')
        employe.prenom = request.POST.get('prenom')
        employe.save()
        messages.success(request, f'Employé {employe.prenom} {employe.nom} modifié avec succès!')
    return redirect('equipe')

def supprimer_employe_pet(request, id_equipe):
    """Vue pour supprimer un employé PET"""
    employe = get_object_or_404(EquipePET, id_equipe=id_equipe)
    if request.method == 'POST':
        nom_complet = f"{employe.prenom} {employe.nom}"
        employe.delete()
        messages.success(request, f'Employé {nom_complet} supprimé avec succès!')
    return redirect('equipe')


from django.db import transaction

# ===== AJOUTER ARRET CARTON =====
# Alternative avec bulk_create pour de meilleures performances
def ajouter_arret_carton(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                arrets_a_creer = []
                
                # Données partagées
                date_base = request.POST.get('date_arret_1')
                ligne_base = request.POST.get('ligne_1')
                equipe_base = request.POST.get('equipe_1')
                quart_base = request.POST.get('quart_1')
                
                for i in range(1, 9):
                    type_arret = request.POST.get(f'type_{i}')
                    machine = request.POST.get(f'machine_{i}')
                    duree_arret = request.POST.get(f'duree_arret_{i}')
                    cause = request.POST.get(f'cause_{i}')
                    responsable = request.POST.get(f'responsable_{i}')
                    
                    if all([type_arret, machine, duree_arret, cause, responsable]):
                        date_arret = request.POST.get(f'date_arret_{i}') or date_base
                        ligne = request.POST.get(f'ligne_{i}') or ligne_base
                        equipe = request.POST.get(f'equipe_{i}') or equipe_base
                        quart = request.POST.get(f'quart_{i}') or quart_base
                        
                        arret = Arret(
                            date_arret=date_arret,
                            ligne_id=ligne,
                            equipe_id=equipe,
                            quart=quart,
                            type_id=type_arret,
                            machine_id=machine,
                            duree_arret=int(duree_arret),
                            cause_id=cause,
                            responsable_id=responsable
                        )
                        arrets_a_creer.append(arret)
                
                # Insertion en bulk (plus rapide)
                if arrets_a_creer:
                    Arret.objects.bulk_create(arrets_a_creer)
                    messages.success(request, f'{len(arrets_a_creer)} arrêt(s) carton ajouté(s) avec succès!')
                else:
                    messages.warning(request, 'Aucun arrêt n\'a été créé.')
                
                return redirect('carton')
                
        except Exception as e:
            print(f"Erreur bulk_create: {str(e)}")
            messages.error(request, f'Erreur lors de l\'ajout: {str(e)}')
            return redirect('carton')

    
    # Reste de votre code pour GET...

def modifier_arret_carton(request, id_arret):
    arret = get_object_or_404(Arret, id_arret=id_arret)
    if request.method == 'POST':
        arret.date_arret = request.POST['date_arret']
        arret.ligne_id = request.POST['ligne']
        arret.equipe_id = request.POST['equipe']
        arret.quart = request.POST['quart']
        arret.type_id = request.POST['type']
        arret.machine_id = request.POST['machine']
        arret.duree_arret = request.POST['duree_arret']
        arret.cause_id = request.POST['cause']
        arret.responsable_id = request.POST['responsable']
        arret.save()
    return redirect('carton')

def supprimer_arret_carton(request, id_arret):
    arret = get_object_or_404(Arret, id_arret=id_arret)
    arret.delete()
    return redirect('carton')
# ============= MÉTHODE 1: Vues basées sur les classes (CBV) =============

# Dans views.py
from django.views.generic import ListView
from .models import Production  # remplacez par votre modèle

class ProductionListView(ListView):
    model = Production
    template_name = 'production_list.html'
    context_object_name = 'productions'
    ordering = ['-id']  # Optionnel: ordre d'affichage
    def get_paginate_by(self, queryset):
        """Récupère le nombre d'éléments par page depuis l'URL"""
        per_page = self.request.GET.get('per_page', 10)
        try:
            per_page = int(per_page)
            # Autoriser différentes valeurs
            if per_page in [3, 5, 10, 25, 50, 100]:
                return per_page
            else:
                return 10
        except ValueError:
            return 10
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Vous pouvez ajouter du contexte supplémentaire ici si nécessaire
        return context



def production_view(request):
    productions = Production.objects.all().order_by('-date')
    lignes = Ligne.objects.all()
    equipe = Equipe.objects.all()
  
    
    # Pagination
    per_page = request.GET.get('per_page', 10)
    paginator = Paginator(productions, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'productions': page_obj,
        'lignes': lignes,
        'equipe': equipe,
        'request': request,

    }
    return render(request, 'rouibapp/production.html', context)


def get_duree_arret_carton(request):
    ligne_id = request.GET.get('ligne_id')
    date = request.GET.get('date')
    quart = request.GET.get('quart')
    # Correction des noms de champs !
    arrets = Arret.objects.filter(ligne_id=ligne_id, date_arret=date, quart=quart)
    duree = arrets.aggregate(models.Sum('duree_arret'))['duree_arret__sum'] or 0
    return JsonResponse({'duree_arret': duree})


from django.db.models import Sum, F, FloatField, ExpressionWrapper

def get_durees_arret_types(request):
    ligne_id = request.GET.get('ligne_id')
    date = request.GET.get('date')
    quart = request.GET.get('quart')
    types = ['Intrants', 'Technique', 'Operationnel', 'Siroperie', 'Utilite', 'Espace']
    result = {}
    for t in types:
        # Convertir duree_arret en float pour l'agrégation
        arrets = Arret.objects.filter(
            ligne_id=ligne_id,
            date_arret=date,
            quart=quart,
            type__nom_type=t
        ).annotate(
            duree_arret_float=ExpressionWrapper(
                F('duree_arret'), output_field=FloatField()
            )
        )
        duree = arrets.aggregate(total=Sum('duree_arret_float'))['total'] or 0
        result[t.lower()] = duree / 60
    return JsonResponse(result)




from .models import Dm
def ajouter_dm(request):
    if request.method == 'POST':
        date_dm = request.POST.get('date_dm')
        ligne_id = request.POST.get('ligne')
        ligne = Ligne.objects.get(pk=ligne_id) if ligne_id else None

        # ✅ Récupérer les arrêts de la même date et ligne
        total_duree = 0
        if date_dm and ligne:
            arrets = Arret.objects.filter(date_arret=date_dm, ligne=ligne)
            # Comme duree_arret est un CharField, il faut convertir en int ou float
            for arret in arrets:
                try:
                    total_duree += float(arret.duree_arret)
                except ValueError:
                    pass  # si la valeur n'est pas un nombre, on ignore

        # Calcul du Prevu = somme des "theorique" de Production pour la même date et ligne
        prevu = 0
        if date_dm and ligne:
            from django.db.models import Sum
            prevu = Production.objects.filter(date=date_dm, ligne=ligne).aggregate(total=Sum('theorique'))['total'] or 0

        # ✅ Création du DM avec le temps d'arrêt calculé
        Dm.objects.create(
            date_dm=date_dm,
            ligne=ligne,
            nbr_arret=request.POST.get('nbr_arret'),
            # On met directement le total dans duree_arret ou crée un nouveau champ temps_arret
            duree_arret = request.POST.get('duree_arret'),
            # ... puis enregistrer dans le modèle DM
            temps_requis=total_duree,  # par ex tu stockes ici le total
            siroperie=safe_float(request.POST.get('siroperie')),
            utilite=safe_float(request.POST.get('utilite')),  # ou utilite selon ton modèle
            MTBF=request.POST.get('MTBF'),
            MTTR=request.POST.get('MTTR'),
            DS=request.POST.get('DS'),
            DE=request.POST.get('DE'),
            DM=request.POST.get('DM'),
            taux_panne=request.POST.get('taux_panne'),
            TRS=request.POST.get('TRS'),
            TRG=request.POST.get('TRG'),
            perte_pourcentage=request.POST.get('perte_pourcentage'),
            TU=request.POST.get('TU'),
            TUBIS=request.POST.get('TUBIS'),
            TRP=request.POST.get('TRP'),
            TRL=request.POST.get('TRL'),
            TRO=request.POST.get('TRO'),
            TRE=request.POST.get('TRE'),
            perte=request.POST.get('perte'),
            realise=request.POST.get('realise'),
            cadence=request.POST.get('cadence'),
            Prevu=prevu,
            HK=request.POST.get('HK'),
        )
        return redirect('dm')
    return redirect('dm')


def get_prevu(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    prevu = 0
    if date and ligne_id:
        try:
            ligne = Ligne.objects.get(pk=ligne_id)
            prevu = Production.objects.filter(date=date, ligne=ligne).aggregate(total=Sum('theorique'))['total'] or 0
        except Ligne.DoesNotExist:
            prevu = 0
    return JsonResponse({'prevu': prevu})


def get_prevu_dm_pet(request):
    date_dm = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    prevu = 0
    if date_dm and ligne_id:
        try:
            # Attention : le champ de date dans ProductionPET est 'date'
            prevu = ProductionPET.objects.filter(date=date_dm, ligne_id=ligne_id).aggregate(total=Sum('theorique'))['total'] or 0
        except Exception:
            prevu = 0
    # On retourne une chaîne pour être compatible avec le CharField de DmPET
    return JsonResponse({'prevu': str(prevu)})


def get_temps_arret(request):
    date = request.GET.get("date")
    ligne = request.GET.get("ligne")

    if date and ligne:
        try:
            # Récupération de la somme des 3 shifts
            arrets = Arret.objects.filter(date_arret=date, ligne_id=ligne)
            duree_totale = 0
            for arret in arrets:
                try:
                    duree = float(arret.duree_arret)
                except Exception:
                    duree = 0
                duree_totale += duree
        except Exception:
            duree_totale = 0
    else:
        duree_totale = 0

    return JsonResponse({"temps_arret": duree_totale})



def get_duree_arret_programme(request):
    date = request.GET.get("date")
    ligne = request.GET.get("ligne")
    total = 0
    if date and ligne:
        # On récupère toutes les productions du jour et de la ligne, tous quarts confondus
        productions = Production.objects.filter(date=date, ligne_id=ligne)
        # On additionne tous les temps_non_programme
        total = sum([p.temps_non_programme or 0 for p in productions]) * 60
    return JsonResponse({"duree_arret_programme": total})



def get_perte_pourcentage(request):
    date = request.GET.get("date")
    ligne = request.GET.get("ligne")
    perte_total = 0
    realise_total = 0
    if date and ligne:
        productions = Production.objects.filter(date=date, ligne_id=ligne)
        for prod in productions:
            perte_total += prod.perte or 0
            realise_total += prod.realise or 0
    pourcentage = (perte_total / realise_total) if realise_total > 0 else 0
    return JsonResponse({"perte_pourcentage": round(pourcentage, 3)})


def get_hk(request):
    date = request.GET.get("date")
    ligne = request.GET.get("ligne")
    total_realise = 0
    if date and ligne:
        productions = Production.objects.filter(date=date, ligne_id=ligne)
        for prod in productions:
            total_realise += prod.realise or 0
    hk = (total_realise / 5) / 100 if total_realise > 0 else 0
    return JsonResponse({"hk": round(hk, 3)})

def get_trs(request):
    date = request.GET.get("date")
    ligne = request.GET.get("ligne")
    total_realise = 0
    total_theorique = 0
    if date and ligne:
        productions = Production.objects.filter(date=date, ligne_id=ligne)
        for prod in productions:
            total_realise += prod.realise or 0
            total_theorique += prod.theorique or 0
    trs = (total_realise / total_theorique * 100) if total_theorique > 0 else 0
    return JsonResponse({"trs": round(trs, 3)})



def get_equipe_by_date_quart_ligne(request):
    date = request.GET.get('date')
    quart = request.GET.get('quart')
    ligne_id = request.GET.get('ligne_id')
    arret = Arret.objects.filter(date_arret=date, quart=quart, ligne_id=ligne_id).first()
    if arret and arret.equipe:
        return JsonResponse({'equipe_id': arret.equipe.id_equipe})
    return JsonResponse({'equipe_id': None})



def get_siroperie(request):
    date = request.GET.get("date")
    ligne = request.GET.get("ligne")
    total = 0
    if date and ligne:
        try:
            type_obj = Type.objects.filter(nom_type="Siroperie").first()
            if type_obj:
                arrets = Arret.objects.filter(date_arret=date, ligne_id=ligne, type=type_obj)
                for arret in arrets:
                    try:
                        total += float(arret.duree_arret)
                    except (ValueError, TypeError):
                        continue
        except Exception:
            total = 0
    return JsonResponse({"siroperie": total})


def get_utilite(request):
    date = request.GET.get("date")
    ligne = request.GET.get("ligne")
    total = 0
    if date and ligne:
        try:
            type_obj = Type.objects.filter(nom_type="Utilite").first()
            if type_obj:
                arrets = Arret.objects.filter(date_arret=date, ligne_id=ligne, type=type_obj)
                for arret in arrets:
                    try:
                        total += float(arret.duree_arret)
                    except (ValueError, TypeError):
                        continue
        except Exception:
            total = 0
    return JsonResponse({"utilite": total})



from django.db.models import Sum

def get_perte(request):
    date = request.GET.get('date')
    ligne = request.GET.get('ligne')
    total_perte = Production.objects.filter(date=date, ligne_id=ligne).aggregate(total=Sum('perte'))['total'] or 0
    return JsonResponse({'perte': total_perte})


def get_realise(request):
    date = request.GET.get('date')
    ligne = request.GET.get('ligne')
    print("DEBUG get_realise:", date, ligne)
    total_realise = Production.objects.filter(date=date, ligne_id=ligne).aggregate(total=models.Sum('realise'))['total'] or 0
    return JsonResponse({'realise': total_realise})


def get_cadence(request):
    date = request.GET.get('date')
    ligne = request.GET.get('ligne')
    # Correction : moyenne de la cadence pour la date et la ligne
    cadence = Production.objects.filter(date=date, ligne_id=ligne).aggregate(avg=models.Avg('cadence'))['avg'] or 0
    return JsonResponse({'cadence': cadence})



@csrf_exempt
@require_http_methods(["POST"])
def supprimer_dm(request, id_dm):
    try:
        # Utiliser id_dm directement depuis l'URL, pas depuis POST
        dm = Dm.objects.get(id_dm=id_dm)
        dm.delete()
        return JsonResponse({'success': True})
    except Dm.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'DM introuvable'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def modifier_dm(request, id_dm):
    if request.method == 'POST':
        try:
            dm = Dm.objects.get(id_dm=id_dm)  # ← Utilise l'ID de l'URL
            # Mets à jour les champs du DM avec les valeurs du formulaire
            dm.date_dm = request.POST.get('date_dm')
            ligne_id = request.POST.get('ligne')
            if ligne_id:
                dm.ligne = Ligne.objects.get(id_ligne=ligne_id)
            dm.duree_arret = request.POST.get('duree_arret')  # ← Correction ici
            dm.nbr_arret = request.POST.get('nbr_arret')
            dm.duree_arret_programme = request.POST.get('duree_arret_programme')
            dm.temps_requis = request.POST.get('temps_requis')
            dm.siroperie = request.POST.get('siroperie')
            dm.utilite = request.POST.get('utilite')
            dm.MTBF = request.POST.get('MTBF')
            dm.MTTR = request.POST.get('MTTR')
            dm.DS = request.POST.get('DS')
            dm.DE = request.POST.get('DE')
            dm.DM = request.POST.get('DM')
            dm.taux_panne = request.POST.get('taux_panne')
            dm.TRS = request.POST.get('TRS')
            dm.TRG = request.POST.get('TRG')
            dm.perte_pourcentage = request.POST.get('perte_pourcentage')
            dm.TU = request.POST.get('TU')
            dm.TUBIS = request.POST.get('TUBIS')
            dm.TRP = request.POST.get('TRP')
            dm.TRL = request.POST.get('TRL')
            dm.TRO = request.POST.get('TRO')
            dm.TRE = request.POST.get('TRE')
            dm.perte = request.POST.get('perte')
            dm.realise = request.POST.get('realise')
            dm.Prevu = request.POST.get('Prevu')
            dm.cadence = request.POST.get('cadence')
            dm.HK = request.POST.get('HK')
            dm.save()
            return redirect('dm')
        except Dm.DoesNotExist:
            return redirect('dm')  # ← Même en cas d'erreur, redirige
    return redirect('dm')






# ===== VUE CORRIGÉE POUR AJOUTER ARRÊT PET =====
def ajouter_arret_pet(request):  # ← Nom cohérent avec l'URL du template
    """Vue pour ajouter un nouvel arrêt PET"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                arrets_a_creer = []
                
                # Données partagées
                date_base = request.POST.get('date_arret_1')
                ligne_base = request.POST.get('ligne_1')
                equipe_base = request.POST.get('equipe_1')
                quart_base = request.POST.get('quart_1')
                
                for i in range(1, 9):
                    type_arret = request.POST.get(f'type_{i}')
                    machine = request.POST.get(f'machine_{i}')
                    duree_arret = request.POST.get(f'duree_arret_{i}')
                    cause = request.POST.get(f'cause_{i}')
                    responsable = request.POST.get(f'responsable_{i}')
                    
                    if all([type_arret, machine, duree_arret, cause, responsable]):
                        date_arret = request.POST.get(f'date_arret_{i}') or date_base
                        ligne = request.POST.get(f'ligne_{i}') or ligne_base
                        equipe = request.POST.get(f'equipe_{i}') or equipe_base
                        quart = request.POST.get(f'quart_{i}') or quart_base
                        
                        arret = ArretPET(
                            date_arret=date_arret,
                            ligne_id=ligne,
                            equipe_id=equipe,
                            quart=quart,
                            type_id=type_arret,
                            machine_id=machine,
                            duree_arret=int(duree_arret),
                            cause_id=cause,
                            responsable_id=responsable
                        )
                        arrets_a_creer.append(arret)
                
                # Insertion en bulk (plus rapide)
                if arrets_a_creer:
                    ArretPET.objects.bulk_create(arrets_a_creer)
                    messages.success(request, f'{len(arrets_a_creer)} arrêt(s) PET ajouté(s) avec succès!')
                else:
                    messages.warning(request, 'Aucun arrêt n\'a été créé.')
                return redirect('PET')
                
        except Exception as e:
            print(f"Erreur bulk_create: {str(e)}")
            messages.error(request, f'Erreur lors de l\'ajout: {str(e)}')
    
    return redirect('PET')


# ===== VUE PET CORRIGÉE =====
def PET(request):
    """Vue pour afficher la page PET avec tous les contextes nécessaires"""
    arrets = ArretPET.objects.select_related(
        'ligne', 'equipe', 'type', 'machine', 'cause', 'responsable'
    ).all().order_by('-date_arret')
    
    context = {
        'arrets': arrets,
        
        # ✅ CORRECTION: Utiliser les bons modèles PET
        'equipes': EquipePET.objects.all(),       # ← EquipePET, pas Equipe
        'lignes': LignePET.objects.all(),         # ← LignePET, pas Ligne  
        'machines': MachinePET.objects.all(),     # ← MachinePET, pas Machine
        'causes': CausesPET.objects.all(),        # ← CausesPET, pas Causes
        'responsables': ResponsablePET.objects.all(), # ← ResponsablePET, pas Responsable
        
        # Type est partagé entre Carton et PET
        'types': Type.objects.all(),
    }
    return render(request, 'rouibapp/PET_arret.html', context)


# ===== FONCTION DE MODIFICATION CORRIGÉE =====
def modifier_arret_pet(request, id_arret_pet):  # ← Nom cohérent avec le template
    """Vue pour modifier un arrêt PET existant"""
    arret = get_object_or_404(ArretPET, id_arretPET=id_arret_pet)
    
    if request.method == 'POST':
        try:
            # ✅ CORRECTION: Utiliser les bons noms de champs
            arret.date_arret = request.POST['date_arret']  # Le modal de modification utilise "date_arret"
            arret.ligne_id = request.POST['ligne']
            arret.equipe_id = request.POST['equipe']
            arret.quart = request.POST['quart']
            arret.type_id = request.POST['type']
            arret.machine_id = request.POST['machine']
            arret.duree_arret = request.POST['duree_arret']
            arret.cause_id = request.POST['cause']
            arret.responsable_id = request.POST['responsable']
            arret.save()
            
            messages.success(request, 'Arrêt PET modifié avec succès!')
            return redirect('PET')
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
            print(f"❌ Erreur modification: {str(e)}")
            
    return redirect('PET')


def supprimer_arret_pet(request, id_arret_pet):  # ← Nom cohérent avec le template
    """Vue pour supprimer un arrêt PET"""
    arret = get_object_or_404(ArretPET, id_arretPET=id_arret_pet)
    
    try:
        arret.delete()
        messages.success(request, 'Arrêt PET supprimé avec succès!')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('PET')






# ===== VIEWS.PY - VUES POUR PRODUCTION PET =====

def production_pet(request):
    """Vue principale pour afficher les productions PET"""
    productions = ProductionPET.objects.select_related(
        'equipe', 'ligne', 'faction'
    ).all().order_by('-date')
    
    # Pagination
    per_page = request.GET.get('per_page', 10)
    paginator = Paginator(productions, per_page)
    page_number = request.GET.get('page')
    productions_page = paginator.get_page(page_number)
    
    context = {
        'productions': productions_page,
        'equipes': EquipePET.objects.all(),
        'lignes': LignePET.objects.all(),
        'factions': FactionPET.objects.all(),
    }
    return render(request, 'rouibapp/production_pet.html', context)


def ajouter_production_pet(request):
    """Vue pour ajouter une nouvelle production PET"""
    if request.method == 'POST':
        try:
            # Récupération et validation des données
            date = request.POST.get('date')
            ligne_id = request.POST.get('ligne')
            equipe_id = request.POST.get('equipe')
            faction_id = request.POST.get('faction')

            # Récupère le nom du quart à partir de l'id de faction
            quart = None
            if faction_id:
                faction_obj = FactionPET.objects.filter(id_faction=faction_id).first()
                quart = faction_obj.nom_faction if faction_obj else None

            # Helpers
            def safe_float(val, default=0.0):
                try:
                    return float(val) if val else default
                except (TypeError, ValueError):
                    return default

            def safe_int(val, default=0):
                try:
                    return int(val) if val else default
                except (TypeError, ValueError):
                    return default

            # Récupération des champs du formulaire
            format_produit = request.POST.get('format_produit', '')
            cadence = safe_float(request.POST.get('cadence'))
            temps_brut = safe_int(request.POST.get('temps_brut'), 8)
            temps_programme = safe_int(request.POST.get('temps_programme'))
            theorique = safe_int(request.POST.get('theorique'))
            realise = safe_int(request.POST.get('realise'))

            # Arrêts programmés
            changement_format_theorique = safe_float(request.POST.get('changement_format_theorique'))
            changement_format_reel = safe_float(request.POST.get('changement_format_reel'))
            entretien_hebdo_theorique = safe_float(request.POST.get('entretien_hebdo_theorique'))
            entretien_hebdo_reel = safe_float(request.POST.get('entretien_hebdo_reel'))
            nep_complet_mp_theorique = safe_float(request.POST.get('nep_complet_mp_theorique'))
            nep_complet_mp_reel = safe_float(request.POST.get('nep_complet_mp_reel'))
            nep_soude_mp_theorique = safe_float(request.POST.get('nep_soude_mp_theorique'))
            nep_soude_mp_reel = safe_float(request.POST.get('nep_soude_mp_reel'))
            nia_theorique = safe_float(request.POST.get('nia_theorique'))
            nia_reel = safe_float(request.POST.get('nia_reel'))
            revision_theorique = safe_float(request.POST.get('revision_theorique'))
            revision_reel = safe_float(request.POST.get('revision_reel'))

            # Pertes
            pertes_vides = safe_int(request.POST.get('pertes_vides'))
            pertes_pleines = safe_int(request.POST.get('pertes_pleines'))

            # Variables de coût
            c = safe_float(request.POST.get('c'))
            # Pour les arrêts non programmés, tu peux utiliser get_duree_type si tu veux calculer à partir des arrêts
            def get_duree_type(type_nom, date, quart):
                type_obj = Type.objects.filter(nom_type=type_nom).first()
                if type_obj and date and quart:
                    total = ArretPET.objects.filter(
                        date_arret=date,
                        quart=quart,
                        type=type_obj
                    ).aggregate(models.Sum('duree_arret'))['duree_arret__sum'] or 0
                    return float(total)
                return 0

            intrants = get_duree_type('Intrants', date, quart)
            technique = get_duree_type('Technique', date, quart)
            operationnel = get_duree_type('Opérationnel', date, quart)
            siroperie = get_duree_type('Siroperie', date, quart)
            utilite = get_duree_type('Utilite', date, quart)
            espace = get_duree_type('Espace', date, quart)
            transports = get_duree_type('Transports', date, quart)

            # === CALCULS AUTOMATIQUES ===
            temps_non_programme = temps_brut - temps_programme
            if cadence > 0 and temps_programme > 0:
                theorique = cadence * temps_programme
            ecart = theorique - realise
            trs_pourcentage = (realise / theorique * 100) if theorique > 0 else 0
            temps_productif = (realise / cadence) if cadence > 0 else 0

            # Écarts
            changement_format_ecart = changement_format_reel - changement_format_theorique
            entretien_hebdo_ecart = entretien_hebdo_reel - entretien_hebdo_theorique
            nep_complet_ecart = nep_complet_mp_reel - nep_complet_mp_theorique
            nep_soude_mp_ecart = nep_soude_mp_reel - nep_soude_mp_theorique
            nia_ecart = nia_reel - nia_theorique
            revision_ecart = revision_reel - revision_theorique

            # Totaux arrêts programmés
            total_arrets_theorique_program = (
                changement_format_theorique +
                entretien_hebdo_theorique +
                nep_complet_mp_theorique +
                nep_soude_mp_theorique +
                nia_theorique +
                revision_theorique
            )
            total_arrets_reels_program = (
                changement_format_reel +
                entretien_hebdo_reel +
                nep_complet_mp_reel +
                nep_soude_mp_reel +
                nia_reel +
                revision_reel
            )
            total_depassements = total_arrets_reels_program - total_arrets_theorique_program

            # Total arrêts non programmés
            total_arrets_non_progr = (
                c + intrants + technique + operationnel +
                siroperie + utilite + espace + transports
            )

            # Total arrêts
            total_arrets = total_arrets_reels_program + total_arrets_non_progr

            # Heures d'arrêts
            heures_arrets = total_arrets / 60 if total_arrets > 0 else 0

            # Temps productif total
            temps_productif_total = temps_brut - total_arrets

            # Total final
            total_final = temps_productif_total

            # Création de la production PET
            ProductionPET.objects.create(
                date=date,
                ligne_id=ligne_id if ligne_id else None,
                equipe_id=equipe_id if equipe_id else None,
                faction_id=faction_id if faction_id else None,
                format_produit=format_produit,
                cadence=cadence,
                temps_brut=temps_brut,
                temps_programme=temps_programme,
                temps_non_programme=temps_non_programme,
                theorique=theorique,
                realise=realise,
                ecart=ecart,
                trs_pourcentage=trs_pourcentage,
                temps_productif=temps_productif,
                heures_arrets=heures_arrets,
                pertes_vides=pertes_vides,
                pertes_pleines=pertes_pleines,
                changement_format_theorique=changement_format_theorique,
                changement_format_reel=changement_format_reel,
                changement_format_ecart=changement_format_ecart,
                entretien_hebdo_theorique=entretien_hebdo_theorique,
                entretien_hebdo_reel=entretien_hebdo_reel,
                entretien_hebdo_ecart=entretien_hebdo_ecart,
                nep_complet_mp_theorique=nep_complet_mp_theorique,
                nep_complet_mp_reel=nep_complet_mp_reel,
                nep_complet_ecart=nep_complet_ecart,
                nep_soude_mp_theorique=nep_soude_mp_theorique,
                nep_soude_mp_reel=nep_soude_mp_reel,
                nep_soude_mp_ecart=nep_soude_mp_ecart,
                nia_theorique=nia_theorique,
                nia_reel=nia_reel,
                nia_ecart=nia_ecart,
                revision_theorique=revision_theorique,
                revision_reel=revision_reel,
                revision_ecart=revision_ecart,
                total_arrets_theorique_program=total_arrets_theorique_program,
                total_arrets_reels_program=total_arrets_reels_program,
                total_depassements=total_depassements,
                total_arrets_non_progr=total_arrets_non_progr,
                total_arrets=total_arrets,
                c=c,
                intrants=intrants,
                technique=technique,
                operationnel=operationnel,
                siroperie=siroperie,
                utilite=utilite,
                espace=espace,
                transports=transports,
                temps_productif_total=temps_productif_total,
                total_final=total_final,
            )
            messages.success(request, 'Production PET ajoutée avec succès!')
            return redirect('production_pet')
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'ajout: {str(e)}')
            return redirect('production_pet')
    return redirect('production_pet')
 


def get_equipe_by_date_quart_ligne_pet(request):
    date = request.GET.get('date')
    faction_id = request.GET.get('faction')
    ligne_id = request.GET.get('ligne_id')

    # Récupérer le nom du quart à partir de la faction
    try:
        faction_obj = FactionPET.objects.get(id_faction=faction_id)
        quart = faction_obj.nom_faction  # correspondance par le nom
    except FactionPET.DoesNotExist:
        return JsonResponse({'equipe_id': None})

    arret = ArretPET.objects.filter(date_arret=date, quart=quart, ligne_id=ligne_id).first()
    if arret and arret.equipe:
        return JsonResponse({'equipe_id': arret.equipe.id_equipe})
    return JsonResponse({'equipe_id': None})


def modifier_production_pet(request, id_production):
    """Vue pour modifier une production PET existante"""
    production = get_object_or_404(ProductionPET, id_production=id_production)
    
    if request.method == 'POST':
        try:
            # Mise à jour des champs principaux
            production.date = request.POST.get('date')
            production.ligne_id = request.POST.get('ligne') if request.POST.get('ligne') else None
            production.equipe_id = request.POST.get('equipe') if request.POST.get('equipe') else None
            production.faction_id = request.POST.get('faction') if request.POST.get('faction') else None
            
            # Fonction helper
            def safe_float(val, default=0.0):
                try:
                    return float(val) if val else default
                except (TypeError, ValueError):
                    return default
            
            def safe_int(val, default=0):
                try:
                    return int(val) if val else default
                except (TypeError, ValueError):
                    return default
            
            # Mise à jour des champs de base
            production.format_produit = request.POST.get('format_produit', '')
            production.cadence = safe_float(request.POST.get('cadence'))
            production.temps_brut = safe_int(request.POST.get('temps_brut'))
            production.temps_programme = safe_int(request.POST.get('temps_programme'))
            production.theorique = int(float(request.POST.get('theorique') or 0))
            production.realise = safe_int(request.POST.get('realise'))
            
            # Personnel
            production.chef_ligne = request.POST.get('chef_ligne', '')
            production.conducteur_souffleuse = request.POST.get('conducteur_souffleuse', '')
            production.conducteur_remplisseuse = request.POST.get('conducteur_remplisseuse', '')
            production.conducteur_etiqueteuse = request.POST.get('conducteur_etiqueteuse', '')
            production.conducteur_fardeleuse = request.POST.get('conducteur_fardeleuse', '')
            production.conducteur_robot = request.POST.get('conducteur_robot', '')
            production.operateur = request.POST.get('operateur', '')
            production.mireur_ligne = request.POST.get('mireur_ligne', '')
            production.cariste = request.POST.get('cariste', '')
            
            # Pertes
            production.pertes_vides = safe_int(request.POST.get('pertes_vides'))
            production.pertes_pleines = safe_int(request.POST.get('pertes_pleines'))
            
            # Variables de coût
            production.c = safe_float(request.POST.get('c'))
            production.intrants = safe_float(request.POST.get('intrants'))
            production.technique = safe_float(request.POST.get('technique'))
            production.operationnel = safe_float(request.POST.get('operationnel'))
            production.siroperie = safe_float(request.POST.get('siroperie'))
            production.utilite = safe_float(request.POST.get('utilite'))
            production.espace = safe_float(request.POST.get('espace'))
            production.transports = safe_float(request.POST.get('transports'))
            
            # Arrêts programmés - Changement de format
            production.changement_format_theorique = safe_float(request.POST.get('changement_format_theorique'))
            production.changement_format_reel = safe_float(request.POST.get('changement_format_reel'))
            
            # Arrêts programmés - Entretien hebdomadaire
            production.entretien_hebdo_theorique = safe_float(request.POST.get('entretien_hebdo_theorique'))
            production.entretien_hebdo_reel = safe_float(request.POST.get('entretien_hebdo_reel'))
            
            # Arrêts programmés - NEP Complet
            production.nep_complet_mp_theorique = safe_float(request.POST.get('nep_complet_mp_theorique'))
            production.nep_complet_mp_reel = safe_float(request.POST.get('nep_complet_mp_reel'))
            
            # Arrêts programmés - NEP Soudé
            production.nep_soude_mp_theorique = safe_float(request.POST.get('nep_soude_mp_theorique'))
            production.nep_soude_mp_reel = safe_float(request.POST.get('nep_soude_mp_reel'))
            
            # Arrêts programmés - NIA
            production.nia_theorique = safe_float(request.POST.get('nia_theorique'))
            production.nia_reel = safe_float(request.POST.get('nia_reel'))
            
            # Arrêts programmés - Révision
            production.revision_theorique = safe_float(request.POST.get('revision_theorique'))
            production.revision_reel = safe_float(request.POST.get('revision_reel'))
            
            # Calculs automatiques après mise à jour des valeurs
            # Écarts
            production.changement_format_ecart = production.changement_format_reel - production.changement_format_theorique
            production.entretien_hebdo_ecart = production.entretien_hebdo_reel - production.entretien_hebdo_theorique
            production.nep_complet_ecart = production.nep_complet_mp_reel - production.nep_complet_mp_theorique
            production.nep_soude_mp_ecart = production.nep_soude_mp_reel - production.nep_soude_mp_theorique
            production.nia_ecart = production.nia_reel - production.nia_theorique
            production.revision_ecart = production.revision_reel - production.revision_theorique
            
            # Totaux pertes
            production.total_perte = production.pertes_vides + production.pertes_pleines
            
            # Écart production
            production.ecart = production.realise - production.theorique
            
            # Temps non programmé
            production.temps_non_programme = production.temps_brut - production.temps_programme
            
            # Totaux arrêts
            production.total_arrets_theorique_program = (
                production.changement_format_theorique + 
                production.entretien_hebdo_theorique + 
                production.nep_complet_mp_theorique + 
                production.nep_soude_mp_theorique + 
                production.nia_theorique + 
                production.revision_theorique
            )
            
            production.total_arrets_reels_program = (
                production.changement_format_reel + 
                production.entretien_hebdo_reel + 
                production.nep_complet_mp_reel + 
                production.nep_soude_mp_reel + 
                production.nia_reel + 
                production.revision_reel
            )
            
            production.total_depassements = production.total_arrets_reels_program - production.total_arrets_theorique_program
            
            # Total arrêts non programmés
            production.total_arrets_non_progr = (
                production.c + production.intrants + production.technique + 
                production.operationnel + production.siroperie + 
                production.utilite + production.espace + production.transports
            )
            
            # Total arrêts
            production.total_arrets = production.total_arrets_reels_program + production.total_arrets_non_progr
            
            # Heures d'arrêts
            production.heures_arrets = production.total_arrets / 60 if production.total_arrets > 0 else 0
            
            # Temps productif
            production.temps_productif = production.temps_programme - production.total_arrets_reels_program
            
            # Temps productif total
            production.temps_productif_total = production.temps_brut - production.total_arrets
            
            # TRS (Taux de Rendement Synthétique)
            if production.temps_programme > 0:
                production.trs_pourcentage = (production.temps_productif / production.temps_programme) * 100
            else:
                production.trs_pourcentage = 0
            
            # Total final
            production.total_final = production.temps_productif_total
            
            production.save()
            messages.success(request, 'Production PET modifiée avec succès!')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    return redirect('production_pet')


def supprimer_production_pet(request, id_production):
    """Vue pour supprimer une production PET"""
    production = get_object_or_404(ProductionPET, id_production=id_production)
    
    try:
        production.delete()
        messages.success(request, 'Production PET supprimée avec succès!')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('production_pet')


def dm_pet(request):
    """Vue pour afficher la page DM PET"""
    from .models import LignePET, ArretPET, ProductionPET, DmPET
    
    context = {
        'lignes': LignePET.objects.all(),
        'arrets': ArretPET.objects.all(),
        'productions': ProductionPET.objects.all(),
        'dms': DmPET.objects.all(),
    }
    return render(request, 'rouibapp/dm_pet.html', context)


def ajouter_dm_pet(request):
    """Vue pour ajouter un nouveau DM PET"""
    if request.method == 'POST':
        from .models import DmPET, LignePET, ArretPET
        
        date_dm = request.POST.get('date_dm')
        ligne_id = request.POST.get('ligne')
        ligne = LignePET.objects.get(pk=ligne_id) if ligne_id else None

        # Récupérer les arrêts de la même date et ligne pour PET
        total_duree = 0
        if date_dm and ligne:
            arrets = ArretPET.objects.filter(date_arret=date_dm, ligne=ligne)
            for arret in arrets:
                try:
                    total_duree += float(arret.duree_arret)
                except ValueError:
                    pass

        # Création du DM PET avec le temps d'arrêt calculé
        DmPET.objects.create(
            date_dm=date_dm,
            ligne=ligne,
            nbr_arret=request.POST.get('nbr_arret'),
            duree_arret = request.POST.get('temps_arret'),
            temps_requis=total_duree,
            siroperie=safe_float(request.POST.get('siroperie')),
            utilite=safe_float(request.POST.get('utilite')),
            MTBF=request.POST.get('MTBF'),
            MTTR=request.POST.get('MTTR'),
            DS=request.POST.get('DS'),
            DE=request.POST.get('DE'),
            DM=request.POST.get('DM'),
            taux_panne=request.POST.get('taux_panne'),
            TRS=request.POST.get('TRS'),
            TRG=request.POST.get('TRG'),
            perte_pourcentage=request.POST.get('perte_pourcentage'),
            TU=request.POST.get('TU'),
            TUBIS=request.POST.get('TUBIS'),
            TRP=request.POST.get('TRP'),
            TRL=request.POST.get('TRL'),
            TRO=request.POST.get('TRO'),
            TRE=request.POST.get('TRE'),
            perte=request.POST.get('perte'),
            realise=request.POST.get('realise'),
            cadence=request.POST.get('cadence'),
            Prevu=request.POST.get('Prevu'),
            HK=request.POST.get('HK'),
        )
        return redirect('dm_pet')
    return redirect('dm_pet')


def get_temps_arret_pet(request):
    """Récupérer le temps d'arrêt total pour PET"""
    from .models import ArretPET
    
    date = request.GET.get("date")
    ligne = request.GET.get("ligne")

    if date and ligne:
        try:
            arrets = ArretPET.objects.filter(date_arret=date, ligne_id=ligne)
            duree_totale = 0
            for arret in arrets:
                try:
                    duree = float(arret.duree_arret)
                except Exception:
                    duree = 0
                duree_totale += duree
        except Exception:
            duree_totale = 0
    else:
        duree_totale = 0

    return JsonResponse({"temps_arret": duree_totale})







def get_duree_arret_by_date_quart_pet(request):
    date = request.GET.get('date')
    faction_id = request.GET.get('faction')  # Doit être un ID !
    duree_arret = 0

    # Récupère le nom du quart à partir de l'id de faction
    quart = None
    if faction_id:
        try:
            faction_obj = FactionPET.objects.get(id_faction=faction_id)
            quart = faction_obj.nom_faction
        except FactionPET.DoesNotExist:
            quart = None

    if date and quart:
        arrets = ArretPET.objects.filter(date_arret=date, quart=quart)
        for arret in arrets:
            try:
                duree_arret += float(arret.duree_arret)
            except:
                pass
    return JsonResponse({'duree_arret': duree_arret})


def get_duree_arret_type_pet(request):
    date = request.GET.get('date')
    faction_id = request.GET.get('faction')  # <-- c'est l'ID !
    type_nom = request.GET.get('type')
    duree = 0

    # Convertir l'ID en nom du quart
    quart = None
    if faction_id:
        faction_obj = FactionPET.objects.filter(id_faction=faction_id).first()
        quart = faction_obj.nom_faction if faction_obj else None

    if date and quart and type_nom:
        type_obj = Type.objects.filter(nom_type=type_nom).first()
        if type_obj:
            arrets = ArretPET.objects.filter(
                date_arret=date,
                quart=quart,  # <-- maintenant c'est le nom du quart
                type=type_obj
            )
            total = sum(float(a.duree_arret) for a in arrets if a.duree_arret)
            duree = total
    return JsonResponse({'duree_arret': duree})


def get_lignepet_by_date_quart_pet(request):
    date = request.GET.get('date')
    faction_id = request.GET.get('faction')
    ligne_id = ''
    equipe_id = ''
    equipe_nom = ''
    if date and faction_id:
        faction_obj = FactionPET.objects.filter(id_faction=faction_id).first()
        quart = faction_obj.nom_faction if faction_obj else None
        if quart:
            arret = ArretPET.objects.filter(date_arret=date, quart=quart).first()
            if arret:
                if arret.ligne:
                    ligne_id = str(arret.ligne.id_ligne)
                if arret.equipe:
                    equipe_id = str(arret.equipe.id_equipe)
                    equipe_nom = f"{arret.equipe.nom} {arret.equipe.prenom}"
    return JsonResponse({
        'ligne_id': ligne_id,
        'equipe_id': equipe_id,
        'equipe_nom': equipe_nom,
    })




@csrf_exempt
@require_http_methods(["POST"])
def supprimer_dm_pet(request, id_dm):
    """Vue pour supprimer un DM PET"""
    from .models import DmPET
    
    try:
        dm = DmPET.objects.get(id_dm=id_dm)
        dm.delete()
        return JsonResponse({'success': True})
    except DmPET.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'DM PET introuvable'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def modifier_dm_pet(request, id_dm):
    """Vue pour modifier un DM PET"""
    from .models import DmPET, LignePET

    if request.method == 'POST':
        try:
            dm = DmPET.objects.get(id_dm=id_dm)  # ← Utilise l'id_dm de l'URL !
            # Mise à jour des champs du DM PET
            dm.date_dm = request.POST.get('date_dm')
            ligne_id = request.POST.get('ligne')
            if ligne_id:
                dm.ligne = LignePET.objects.get(id_ligne=ligne_id)
            dm.duree_arret = request.POST.get('temps_arret')
            dm.nbr_arret = request.POST.get('nbr_arret')
            dm.duree_arret_programme = request.POST.get('duree_arret_programme')
            dm.temps_requis = request.POST.get('temps_requis')
            dm.siroperie = request.POST.get('siroperie')
            dm.utilite = request.POST.get('utilite')
            dm.MTBF = request.POST.get('MTBF')
            dm.MTTR = request.POST.get('MTTR')
            dm.DS = request.POST.get('DS')
            dm.DE = request.POST.get('DE')
            dm.DM = request.POST.get('DM')
            dm.taux_panne = request.POST.get('taux_panne')
            dm.TRS = request.POST.get('TRS')
            dm.TRG = request.POST.get('TRG')
            dm.perte_pourcentage = request.POST.get('perte_pourcentage')
            dm.TU = request.POST.get('TU')
            dm.TUBIS = request.POST.get('TUBIS')
            dm.TRP = request.POST.get('TRP')
            dm.TRL = request.POST.get('TRL')
            dm.TRO = request.POST.get('TRO')
            dm.TRE = request.POST.get('TRE')
            dm.perte = request.POST.get('perte')
            dm.realise = request.POST.get('realise')
            dm.Prevu = request.POST.get('Prevu')
            dm.cadence = request.POST.get('cadence')
            dm.HK = request.POST.get('HK')
            dm.save()
            return JsonResponse({'success': True})
        except DmPET.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'DM PET introuvable'})
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


def get_temps_arret_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    temps_arret = 0
    if date and ligne_id:
        arrets = ArretPET.objects.filter(date_arret=date, ligne_id=ligne_id)
        temps_arret = sum(float(a.duree_arret) for a in arrets if a.duree_arret)
    return JsonResponse({'temps_arret': temps_arret})


def get_duree_arret_programme_dm_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    duree_arret_programme = 0
    if date and ligne_id:
        # Filtrer les trois quarts pour la date et la ligne
        prods = ProductionPET.objects.filter(
            date=date,
            ligne_id=ligne_id,
            faction__nom_faction__in=['Matin', 'Soir', 'Nuit']
        )
        # Additionner les temps non programmés
        total_temps_non_programme = sum(float(p.temps_non_programme) for p in prods if p.temps_non_programme)
        duree_arret_programme = total_temps_non_programme * 60
    return JsonResponse({'duree_arret_programme': duree_arret_programme})

def get_siroperie_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    siroperie = 0
    if date and ligne_id:
        type_obj = Type.objects.filter(nom_type="Siroperie").first()
        if type_obj:
            arrets = ArretPET.objects.filter(date_arret=date, ligne_id=ligne_id, type=type_obj)
            siroperie = sum(float(a.duree_arret) for a in arrets if a.duree_arret)
    return JsonResponse({'siroperie': siroperie})

def get_utilite_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    utilite = 0
    if date and ligne_id:
        type_obj = Type.objects.filter(nom_type="Utilite").first()
        if type_obj:
            arrets = ArretPET.objects.filter(date_arret=date, ligne_id=ligne_id, type=type_obj)
            utilite = sum(float(a.duree_arret) for a in arrets if a.duree_arret)
    return JsonResponse({'utilite': utilite})

def get_trg_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    trg = 0
    if date and ligne_id:
        prod = ProductionPET.objects.filter(date=date, ligne_id=ligne_id).first()
        if prod and hasattr(prod, 'TRG'):
            trg = prod.TRG
    return JsonResponse({'trg': trg})

def get_perte_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    perte = 0
    if date and ligne_id:
        prods = ProductionPET.objects.filter(
            date=date,
            ligne_id=ligne_id,
            faction__nom_faction__in=['Matin', 'Soir', 'Nuit']
        )
        perte = sum(
            (float(p.pertes_vides or 0) + float(p.pertes_pleines or 0))
            for p in prods
        )
    return JsonResponse({'perte': perte})

def get_realise_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    realise = 0
    if date and ligne_id:
        prods = ProductionPET.objects.filter(
            date=date,
            ligne_id=ligne_id,
            faction__nom_faction__in=['Matin', 'Soir', 'Nuit']
        )
        realise = sum(float(p.realise) for p in prods if p.realise)
    return JsonResponse({'realise': realise})

def get_perte_pourcentage_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    perte_pourcentage = 0
    if date and ligne_id:
        prod = ProductionPET.objects.filter(date=date, ligne_id=ligne_id).first()
        if prod and hasattr(prod, 'perte_pourcentage'):
            perte_pourcentage = prod.perte_pourcentage
    return JsonResponse({'perte_pourcentage': perte_pourcentage})

def get_hk_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    hk = 0
    if date and ligne_id:
        prod = ProductionPET.objects.filter(date=date, ligne_id=ligne_id).first()
        if prod and hasattr(prod, 'HK'):
            hk = prod.HK
    return JsonResponse({'hk': hk})

def get_cadence_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    cadence = 0
    if date and ligne_id:
        # On récupère la première production PET pour la date et la ligne
        prod = ProductionPET.objects.filter(date=date, ligne_id=ligne_id).first()
        if prod and hasattr(prod, 'cadence'):
            cadence = prod.cadence
    return JsonResponse({'cadence': cadence})

def get_trs_pet(request):
    date = request.GET.get('date')
    ligne_id = request.GET.get('ligne')
    trs = 0
    if date and ligne_id:
        prod = ProductionPET.objects.filter(date=date, ligne_id=ligne_id).first()
        if prod and prod.theorique and prod.realise:
            trs = (float(prod.realise) / float(prod.theorique)) 
    return JsonResponse({'trs': round(trs, 3)})

import openpyxl
from django.http import HttpResponse

def export_arrets_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arrêts Carton"
    ws.append([
        "Date", "Ligne", "Équipe", "Quart", "Type", "Machine", "Durée (min)", "Cause", "Responsable"
    ])
    from .models import Arret
    for arret in Arret.objects.all():
        ws.append([
            arret.date_arret.strftime("%d/%m/%Y"),
            arret.ligne.nom_ligne,
            f"{arret.equipe.nom} {arret.equipe.prenom}",
            arret.quart,
            arret.type.nom_type,
            arret.machine.nom_machine,
            arret.duree_arret,
            arret.cause.nom_cause,
            arret.responsable.nom_resp,
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=arrets_carton.xlsx'
    wb.save(response)
    return response



def export_production_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Carton"
    ws.append([
        "Date", "Ligne", "Équipe", "Quart", "Cadence", "Théorique", "Réalisé", "TRS", "Perte", "Ecart"
    ])
    from .models import Production
    for prod in Production.objects.select_related('ligne', 'equipe').all():
        ws.append([
            prod.date.strftime("%d/%m/%Y") if prod.date else "",
            prod.ligne.nom_ligne if prod.ligne else "",
            f"{prod.equipe.nom} {prod.equipe.prenom}" if prod.equipe else "",
            prod.quart,
            prod.cadence,
            prod.theorique,
            prod.realise,
            prod.trs,
            prod.perte,
            prod.ecart,
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=production_carton.xlsx'
    wb.save(response)
    return response

import openpyxl
from django.http import HttpResponse

def export_dm_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DM Carton"
    ws.append([
        "Date DM", "Ligne", "Nombre d'arrêts", "Durée arrêt", "Temps requis", "Siroperie", "Utilite",
        "MTBF", "MTTR", "DS", "DE", "DM", "Taux de panne", "TRS", "TRG", "Perte %", "TU", "TUBIS",
        "TRP", "TRL", "TRO", "TRE", "Perte", "Réalisé", "Cadence", "Prévu", "HK"
    ])
    from .models import Dm
    for dm in Dm.objects.select_related('ligne').all():
        ws.append([
            dm.date_dm.strftime("%d/%m/%Y") if dm.date_dm else "",
            dm.ligne.nom_ligne if dm.ligne else "",
            dm.nbr_arret,
            dm.duree_arret,
            dm.temps_requis,
            dm.siroperie,
            dm.utilite,
            dm.MTBF,
            dm.MTTR,
            dm.DS,
            dm.DE,
            dm.DM,
            dm.taux_panne,
            dm.TRS,
            dm.TRG,
            dm.perte_pourcentage,
            dm.TU,
            dm.TUBIS,
            dm.TRP,
            dm.TRL,
            dm.TRO,
            dm.TRE,
            dm.perte,
            dm.realise,
            dm.cadence,
            dm.Prevu,
            dm.HK,
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=dm_carton.xlsx'
    wb.save(response)
    return response

def export_arrets_pet_excel(request):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arrêts PET"
    ws.append([
        "Date", "Ligne", "Équipe", "Quart", "Type", "Machine", "Durée (min)", "Cause", "Responsable"
    ])
    from .models import ArretPET
    for arret in ArretPET.objects.select_related('ligne', 'equipe', 'type', 'machine', 'cause', 'responsable').all():
        ws.append([
            arret.date_arret.strftime("%d/%m/%Y") if arret.date_arret else "",
            arret.ligne.nom_ligne if arret.ligne else "",
            f"{arret.equipe.nom} {arret.equipe.prenom}" if arret.equipe else "",
            arret.quart,
            arret.type.nom_type if arret.type else "",
            arret.machine.nom_machine if arret.machine else "",
            arret.duree_arret,
            arret.cause.nom_cause if arret.cause else "",
            arret.responsable.nom_resp if arret.responsable else "",
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=arrets_pet.xlsx'
    wb.save(response)
    return response

def export_production_pet_excel(request):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production PET"
    ws.append([
        "Date", "Ligne", "Équipe", "Quart", "Cadence", "Théorique", "Réalisé", "TRS", "Pertes vides", "Pertes pleines", "Ecart"
    ])
    from .models import ProductionPET, FactionPET
    for prod in ProductionPET.objects.select_related('ligne', 'equipe', 'faction').all():
        ws.append([
            prod.date.strftime("%d/%m/%Y") if prod.date else "",
            prod.ligne.nom_ligne if prod.ligne else "",
            f"{prod.equipe.nom} {prod.equipe.prenom}" if prod.equipe else "",
            prod.faction.nom_faction if prod.faction else "",
            prod.cadence,
            prod.theorique,
            prod.realise,
            prod.trs_pourcentage,
            prod.pertes_vides,
            prod.pertes_pleines,
            prod.ecart if hasattr(prod, 'ecart') else "",
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=production_pet.xlsx'
    wb.save(response)
    return response

def export_dm_pet_excel(request):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DM PET"
    ws.append([
        "Date DM", "Ligne", "Nombre d'arrêts", "Durée arrêt", "Temps requis", "Siroperie", "Utilite",
        "MTBF", "MTTR", "DS", "DE", "DM", "Taux de panne", "TRS", "TRG", "Perte %", "TU", "TUBIS",
        "TRP", "TRL", "TRO", "TRE", "Perte", "Réalisé", "Cadence", "Prévu", "HK"
    ])
    from .models import DmPET
    for dm in DmPET.objects.select_related('ligne').all():
        ws.append([
            dm.date_dm.strftime("%d/%m/%Y") if dm.date_dm else "",
            dm.ligne.nom_ligne if dm.ligne else "",
            dm.nbr_arret,
            dm.duree_arret,
            dm.temps_requis,
            dm.siroperie,
            dm.utilite,
            dm.MTBF,
            dm.MTTR,
            dm.DS,
            dm.DE,
            dm.DM,
            dm.taux_panne,
            dm.TRS,
            dm.TRG,
            dm.perte_pourcentage,
            dm.TU,
            dm.TUBIS,
            dm.TRP,
            dm.TRL,
            dm.TRO,
            dm.TRE,
            dm.perte,
            dm.realise,
            dm.cadence,
            dm.Prevu,
            dm.HK,
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=dm_pet.xlsx'
    wb.save(response)
    return response


# STATISTIQUES DE LA PAGE D'ACCEUIL 

from django.shortcuts import render
from .models import Production, Ligne, Dm, ProductionPET, DmPET

def home(request):
    date = request.GET.get('date')
    lignes_noms = ['F', 'G', 'H', 'N']
    stats = []

    # Carton
    for nom_ligne in lignes_noms:
        ligne_obj = Ligne.objects.filter(nom_ligne=nom_ligne).first()
        dm = Dm.objects.filter(ligne=ligne_obj, date_dm=date).first() if ligne_obj and date else None
        
        stats.append({
            'ligne': nom_ligne,
            'planifie': dm.Prevu if dm and hasattr(dm, 'Prevu') else 0,
            'realise': dm.realise if dm and hasattr(dm, 'realise') else 0,
            'perte': dm.perte if dm and hasattr(dm, 'perte') else 0,  # ← AJOUTE CETTE LIGNE
            'performance': round((dm.realise * 100 / dm.Prevu), 3) if dm and dm.Prevu else 0,
            'trs': round(float(dm.TRS), 3) if dm and dm.TRS else 0,
            'trg': round(float(dm.TRG), 3) if dm and dm.TRG else 0
        })

    return render(request, 'rouibapp/home.html', {
        'stats': stats,
        'date': date
    })

def home_pet(request):
    date = request.GET.get('date')
    lignes_pet = ['PET 1', 'PET 2']
    stats_pet = []

    for nom_ligne_pet in lignes_pet:
        ligne_pet_obj = LignePET.objects.filter(nom_ligne=nom_ligne_pet).first()
        dm_pet = DmPET.objects.filter(ligne=ligne_pet_obj, date_dm=date).first() if ligne_pet_obj and date else None

        # Récupérer la somme des pertes pour la ligne et la date
        perte = 0
        if ligne_pet_obj and date:
            perte = ProductionPET.objects.filter(
                ligne=ligne_pet_obj,
                date=date
            ).aggregate(
                total=Sum('pertes_vides') + Sum('pertes_pleines')
            )['total'] or 0

        stats_pet.append({
            'ligne': nom_ligne_pet,
            'planifie': dm_pet.Prevu if dm_pet and hasattr(dm_pet, 'Prevu') else 0,
            'realise': dm_pet.realise if dm_pet and hasattr(dm_pet, 'realise') else 0,
            'perte': perte,
            'performance': (round((float(dm_pet.realise) * 100 / float(dm_pet.Prevu)), 3)if dm_pet and dm_pet.Prevu and dm_pet.realise else 0),
            'trs': round(float(dm_pet.TRS), 3) if dm_pet and dm_pet.TRS else 0,
            'trg': round(float(dm_pet.TRG), 3) if dm_pet and dm_pet.TRG else 0
        })
    
    return render(request, 'rouibapp/home_pet.html', {
        'stats_pet': stats_pet,
        'date': date
    })


from django.db.models import Sum, Avg

def statistiques(request):
    date = request.GET.get('date')
    equipes = Equipe.objects.all()
    types = Type.objects.all()
    equipe_labels = [e.nom for e in equipes]  # Remplie une seule fois
    realise_values = []
    trs_values = []
    type_labels = []
    arret_values = []
    temps_arret = []
    temps_productif = []
    responsables = Responsable.objects.all()
    responsable_labels = [r.nom_resp for r in responsables]
    arret_grouped_responsable = {r.nom_resp: [] for r in responsables}
    perte_pourcent_values = []

    # Réalisé, TRS, Perte % par équipe
    for equipe in equipes:
        prod_qs = Production.objects.filter(equipe=equipe)
        if date:
            prod_qs = prod_qs.filter(date=date)
        total_realise = prod_qs.aggregate(total=Sum('realise'))['total'] or 0
        avg_trs = prod_qs.aggregate(avg=Avg('trs'))['avg'] or 0
        total_perte = prod_qs.aggregate(total=Sum('perte'))['total'] or 0
        perte_pourcent = (total_perte / total_realise * 100) if total_realise > 0 else 0

        realise_values.append(total_realise)
        trs_values.append(round(avg_trs, 3))
        perte_pourcent_values.append(round(perte_pourcent, 3))
    
    equipe_labels.append("Total")
    realise_values.append(sum(realise_values))
    trs_values.append(round(sum(trs_values) / len(trs_values), 3) if trs_values else 0)
    perte_pourcent_values.append(round(sum(perte_pourcent_values) / len(perte_pourcent_values), 2) if perte_pourcent_values else 0)

    # Temps d'arrêt par type
    for t in types:
        arret_qs = Arret.objects.filter(type=t)
        if date:
            arret_qs = arret_qs.filter(date_arret=date)
        total_arret = arret_qs.aggregate(total=Sum('duree_arret'))['total'] or 0
        type_labels.append(t.nom_type)
        arret_values.append(total_arret)

    type_labels.append("Total")
    arret_values.append(sum(arret_values))    

    # Histogramme empilé temps productif / temps d'arrêt
    TEMPS_PLANIFIE = 480  # exemple: 8h en minutes
    for equipe in equipes:
        arret_qs = Arret.objects.filter(equipe=equipe)
        if date:
            arret_qs = arret_qs.filter(date_arret=date)
        total_arret = arret_qs.aggregate(total=Sum('duree_arret'))['total'] or 0
        temps_arret.append(total_arret)
        temps_productif.append(max(TEMPS_PLANIFIE - total_arret, 0))

    temps_arret.append(sum(temps_arret))
    temps_productif.append(sum(temps_productif))

    # Histogramme groupé temps d'arrêt par équipe et type
    arret_grouped = {t.nom_type: [] for t in types}
    for equipe in equipes:
        for t in types:
            arret_qs = Arret.objects.filter(equipe=equipe, type=t)
            if date:
                arret_qs = arret_qs.filter(date_arret=date)
            total_arret = arret_qs.aggregate(total=Sum('duree_arret'))['total'] or 0
            arret_grouped[t.nom_type].append(total_arret)

    for t in types:
        arret_grouped[t.nom_type].append(sum(arret_grouped[t.nom_type]))

    # Histogramme groupé temps d'arrêt par équipe et responsable
    for equipe in equipes:
        for r in responsables:
            arret_qs = Arret.objects.filter(equipe=equipe, responsable=r)
            if date:
                arret_qs = arret_qs.filter(date_arret=date)
            total_arret = arret_qs.aggregate(total=Sum('duree_arret'))['total'] or 0
            arret_grouped_responsable[r.nom_resp].append(total_arret)

    for r in responsables:
        arret_grouped_responsable[r.nom_resp].append(sum(arret_grouped_responsable[r.nom_resp]))
    responsable_labels.append("Total")

    context = {
        'equipe_labels': equipe_labels,
        'realise_values': realise_values,
        'trs_values': trs_values,
        'date': date,
        'type_labels': type_labels,
        'arret_values': arret_values,
        'temps_arret': temps_arret,
        'temps_productif': temps_productif,
        'arret_grouped': arret_grouped,
        'responsable_labels': responsable_labels,
        'arret_grouped_responsable': arret_grouped_responsable,
        'perte_pourcent_values': perte_pourcent_values,
    }
    return render(request, 'rouibapp/statistiques.html', context)



def statistiques_pet(request):
    date = request.GET.get('date')
    equipes = EquipePET.objects.all()
    types = Type.objects.all()
    responsables = ResponsablePET.objects.all()

    equipe_labels = []
    realise_values = []
    trs_values = []
    perte_pourcent_values = []
    type_labels = []
    arret_values = []
    temps_arret = []
    temps_productif = []
    responsable_labels = [r.nom_resp for r in responsables]
    arret_grouped_responsable = {r.nom_resp: [] for r in responsables}

    # Réalisé, TRS, Perte % par équipe
    for equipe in equipes:
        prod_qs = ProductionPET.objects.filter(equipe=equipe)
        if date:
            prod_qs = prod_qs.filter(date=date)
        total_realise = prod_qs.aggregate(total=Sum('realise'))['total'] or 0
        avg_trs = prod_qs.aggregate(avg=Avg('trs_pourcentage'))['avg'] or 0
        total_perte = prod_qs.aggregate(total=Sum('pertes_vides') + Sum('pertes_pleines'))['total'] or 0
        perte_pourcent = (total_perte / total_realise * 100) if total_realise > 0 else 0

        equipe_labels.append(getattr(equipe, 'nom', f"{equipe.nom} {getattr(equipe, 'prenom', '')}"))
        realise_values.append(total_realise)
        trs_values.append(round(avg_trs, 3))
        perte_pourcent_values.append(round(perte_pourcent, 3))

    # Ajout du total pour les équipes
    equipe_labels.append("Total")
    realise_values.append(sum(realise_values))
    trs_values.append(round(sum(trs_values) / len(trs_values), 3) if trs_values else 0)
    perte_pourcent_values.append(round(sum(perte_pourcent_values) / len(perte_pourcent_values), 3) if perte_pourcent_values else 0)

    # Temps d'arrêt par type
    for t in types:
        arret_qs = ArretPET.objects.filter(type=t)
        if date:
            arret_qs = arret_qs.filter(date_arret=date)
        total_arret = arret_qs.aggregate(total=Sum('duree_arret'))['total'] or 0
        type_labels.append(t.nom_type)
        arret_values.append(total_arret)

    # Ajout du total pour les types
    type_labels.append("Total")
    arret_values.append(sum(arret_values))

    # Histogramme empilé temps productif / temps d'arrêt
    TEMPS_PLANIFIE = 480  # exemple: 8h en minutes
    for equipe in equipes:
        arret_qs = ArretPET.objects.filter(equipe=equipe)
        if date:
            arret_qs = arret_qs.filter(date_arret=date)
        total_arret = arret_qs.aggregate(total=Sum('duree_arret'))['total'] or 0
        temps_arret.append(total_arret)
        temps_productif.append(max(TEMPS_PLANIFIE - total_arret, 0))

    # Ajout du total pour les temps
    temps_arret.append(sum(temps_arret))
    temps_productif.append(sum(temps_productif))

    # Histogramme groupé temps d'arrêt par équipe et type
    arret_grouped = {t.nom_type: [] for t in types}
    for equipe in equipes:
        for t in types:
            arret_qs = ArretPET.objects.filter(equipe=equipe, type=t)
            if date:
                arret_qs = arret_qs.filter(date_arret=date)
            total_arret = arret_qs.aggregate(total=Sum('duree_arret'))['total'] or 0
            arret_grouped[t.nom_type].append(total_arret)
    # Ajout du total pour chaque type
    for t in types:
        arret_grouped[t.nom_type].append(sum(arret_grouped[t.nom_type]))

    # Histogramme groupé temps d'arrêt par équipe et responsable
    for equipe in equipes:
        for r in responsables:
            arret_qs = ArretPET.objects.filter(equipe=equipe, responsable=r)
            if date:
                arret_qs = arret_qs.filter(date_arret=date)
            total_arret = arret_qs.aggregate(total=Sum('duree_arret'))['total'] or 0
            arret_grouped_responsable[r.nom_resp].append(total_arret)
    # Ajout du total pour chaque responsable
    for r in responsables:
        arret_grouped_responsable[r.nom_resp].append(sum(arret_grouped_responsable[r.nom_resp]))
    responsable_labels.append("Total")

    context = {
        'equipe_labels': equipe_labels,
        'realise_values': realise_values,
        'trs_values': trs_values,
        'date': date,
        'type_labels': type_labels,
        'arret_values': arret_values,
        'temps_arret': temps_arret,
        'temps_productif': temps_productif,
        'arret_grouped': arret_grouped,
        'responsable_labels': responsable_labels,
        'arret_grouped_responsable': arret_grouped_responsable,
        'perte_pourcent_values': perte_pourcent_values,
    }
    return render(request, 'rouibapp/statistiques_pet.html', context)